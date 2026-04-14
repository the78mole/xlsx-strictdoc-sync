"""Excel engine: read/write requirements from/to Excel workbooks.

Supports two access modes:

* **Table mode** – targets an Excel *ListObject* (Formatted Table) by its
  display name.  Column names match the table's header row.  New rows are
  appended *inside* the table boundaries and the table reference is expanded
  automatically.

* **Legacy mode** – targets a plain worksheet by its sheet name.  Columns are
  addressed using A1-style letters (``"A"``, ``"B"``, ``"AA"`` …).  Row 1 is
  treated as a header row; data starts from row 2.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries

from .models import Requirement

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table
    from openpyxl.worksheet.worksheet import Worksheet

    from .config_manager import SectionMapping


class ExcelEngineError(Exception):
    """Raised when an Excel operation fails."""


class ExcelEngine:
    """Facade for all openpyxl operations.

    Open the workbook via :meth:`open` (or use the context-manager protocol),
    then call :meth:`read_requirements` / :meth:`write_requirements` for each
    section mapping.  Finally call :meth:`save` and :meth:`close`.

    Example::

        with ExcelEngine("requirements.xlsx") as eng:
            reqs = eng.read_requirements(mapping)
    """

    def __init__(self, excel_path: str | Path) -> None:
        self.excel_path = Path(excel_path)
        self._wb: Workbook | None = None

    # ------------------------------------------------------------------
    # Context-manager support
    # ------------------------------------------------------------------

    def __enter__(self) -> "ExcelEngine":
        self.open()
        return self

    def __exit__(self, *_: object) -> None:
        self.close()

    # ------------------------------------------------------------------
    # Lifecycle
    # ------------------------------------------------------------------

    def open(self) -> None:
        """Open the workbook for reading and writing."""
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
        self._wb = openpyxl.load_workbook(str(self.excel_path))

    def save(self, path: str | Path | None = None) -> None:
        """Save the workbook; defaults to the original path."""
        self._require_open()
        save_path = Path(path) if path else self.excel_path
        self._wb.save(str(save_path))  # type: ignore[union-attr]

    def close(self) -> None:
        """Close the workbook without saving."""
        if self._wb is not None:
            self._wb.close()
            self._wb = None

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def read_requirements(self, mapping: "SectionMapping") -> list[Requirement]:
        """Read all requirements for *mapping* from the open workbook.

        Args:
            mapping: Section configuration describing how to locate and
                interpret the data.

        Returns:
            List of :class:`~.models.Requirement` objects (rows with empty
            UID are silently skipped).

        Raises:
            ExcelEngineError: If the sheet/table named in *mapping* is not
                found.
        """
        self._require_open()
        if mapping.mode == "table":
            return self._read_table_mode(mapping)
        return self._read_legacy_mode(mapping)

    def write_requirements(
        self,
        requirements: list[Requirement],
        mapping: "SectionMapping",
    ) -> None:
        """Write *requirements* back to the Excel workbook.

        Existing rows whose UID is already present in the workbook are
        **updated in-place**.  New requirements are **appended** at the end
        of the table / sheet.

        Args:
            requirements: Requirements to write.
            mapping: Section configuration.

        Raises:
            ExcelEngineError: If the sheet/table is not found.
        """
        self._require_open()
        if mapping.mode == "table":
            self._write_table_mode(requirements, mapping)
        else:
            self._write_legacy_mode(requirements, mapping)

    def list_sheets(self) -> list[str]:
        """Return all sheet names in the workbook."""
        self._require_open()
        return list(self._wb.sheetnames)  # type: ignore[union-attr]

    def list_tables(self) -> dict[str, str]:
        """Return a mapping of table display-name → sheet name."""
        self._require_open()
        result: dict[str, str] = {}
        for ws in self._wb.worksheets:  # type: ignore[union-attr]
            for tbl in ws.tables.values():
                result[tbl.displayName or tbl.name] = ws.title
        return result

    def get_sheet_headers(self, sheet_name: str, header_row: int = 1) -> list[str]:
        """Return non-empty header cell values from *sheet_name*.

        Args:
            sheet_name: Worksheet name.
            header_row: 1-based row index of the header (default ``1``).

        Returns:
            List of header strings (empty cells excluded).
        """
        self._require_open()
        ws: Worksheet = self._require_sheet(sheet_name)
        return [
            str(cell.value)
            for cell in ws[header_row]  # type: ignore[index]
            if cell.value is not None
        ]

    def get_table_headers(self, table_name: str) -> list[str]:
        """Return column header names for a named Excel table."""
        self._require_open()
        ws, tbl = self._find_table(table_name)
        min_col, min_row, max_col, _ = range_boundaries(tbl.ref)
        return [
            str(ws.cell(min_row, c).value or "")
            for c in range(min_col, max_col + 1)
        ]

    # ------------------------------------------------------------------
    # Table-mode internals
    # ------------------------------------------------------------------

    def _read_table_mode(self, mapping: "SectionMapping") -> list[Requirement]:
        ws, tbl = self._find_table(mapping.anchor)
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)

        # Build header → column-index lookup (1-based)
        headers: dict[str, int] = {
            str(ws.cell(min_row, c).value): c
            for c in range(min_col, max_col + 1)
            if ws.cell(min_row, c).value is not None
        }

        uid_cidx = self._col_idx_by_header(headers, mapping.uid_col, mapping.name, "uid_col")

        requirements: list[Requirement] = []
        for row in range(min_row + 1, max_row + 1):
            uid_val = ws.cell(row, uid_cidx).value
            if not uid_val:
                continue
            req = self._build_requirement_from_row(
                ws=ws,
                row=row,
                headers=headers,
                mapping=mapping,
                uid=str(uid_val),
                col_resolver=lambda col_name: headers.get(col_name),
            )
            requirements.append(req)
        return requirements

    def _write_table_mode(
        self,
        requirements: list[Requirement],
        mapping: "SectionMapping",
    ) -> None:
        ws, tbl = self._find_table(mapping.anchor)
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)

        headers: dict[str, int] = {
            str(ws.cell(min_row, c).value): c
            for c in range(min_col, max_col + 1)
            if ws.cell(min_row, c).value is not None
        }

        uid_cidx = self._col_idx_by_header(headers, mapping.uid_col, mapping.name, "uid_col")

        # Build UID → existing row-index map
        uid_row: dict[str, int] = {}
        for row in range(min_row + 1, max_row + 1):
            uid_val = ws.cell(row, uid_cidx).value
            if uid_val:
                uid_row[str(uid_val)] = row

        current_last_row = max_row
        for req in requirements:
            if req.uid in uid_row:
                target_row = uid_row[req.uid]
            else:
                current_last_row += 1
                target_row = current_last_row

            self._write_requirement_to_row(ws, target_row, headers, mapping, req)

        # Expand table reference if rows were added
        if current_last_row > max_row:
            new_ref = (
                f"{get_column_letter(min_col)}{min_row}"
                f":{get_column_letter(max_col)}{current_last_row}"
            )
            tbl.ref = new_ref

    # ------------------------------------------------------------------
    # Legacy-mode internals
    # ------------------------------------------------------------------

    def _read_legacy_mode(self, mapping: "SectionMapping") -> list[Requirement]:
        ws = self._require_sheet(mapping.anchor)

        # In legacy mode, uid_col etc. are A1 column letters
        uid_cidx = column_index_from_string(mapping.uid_col)

        def _legacy_col_resolver(col_letter: str) -> int | None:
            if not col_letter:
                return None
            return column_index_from_string(col_letter)

        requirements: list[Requirement] = []
        for row_idx in range(2, ws.max_row + 1):  # row 1 is header
            uid_val = ws.cell(row_idx, uid_cidx).value
            if not uid_val:
                continue

            req = self._build_requirement_from_row(
                ws=ws,
                row=row_idx,
                headers={},
                mapping=mapping,
                uid=str(uid_val),
                col_resolver=_legacy_col_resolver,
            )
            requirements.append(req)
        return requirements

    def _write_legacy_mode(
        self,
        requirements: list[Requirement],
        mapping: "SectionMapping",
    ) -> None:
        ws = self._require_sheet(mapping.anchor)

        uid_cidx = column_index_from_string(mapping.uid_col)

        # Build UID → row map
        uid_row: dict[str, int] = {}
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row_idx, uid_cidx).value
            if val:
                uid_row[str(val)] = row_idx

        next_new_row = ws.max_row + 1
        for req in requirements:
            if req.uid in uid_row:
                target_row = uid_row[req.uid]
            else:
                target_row = next_new_row
                next_new_row += 1

            def legacy_col_resolver(col_letter: str) -> int | None:
                if not col_letter:
                    return None
                return column_index_from_string(col_letter)

            self._write_requirement_to_row(ws, target_row, {}, mapping, req)

    # ------------------------------------------------------------------
    # Shared row-level helpers
    # ------------------------------------------------------------------

    def _build_requirement_from_row(
        self,
        ws: "Worksheet",
        row: int,
        headers: dict[str, int],
        mapping: "SectionMapping",
        uid: str,
        col_resolver,
    ) -> Requirement:
        title = self._cell_str(ws, row, col_resolver(mapping.title_col))
        statement = self._cell_str(ws, row, col_resolver(mapping.statement_col))
        relations_raw = self._cell_str(ws, row, col_resolver(mapping.relations_col))
        relations = [r.strip() for r in relations_raw.split(";") if r.strip()]

        custom_fields: dict[str, str] = {}
        for excel_col, sdoc_field in mapping.extra_cols.items():
            custom_fields[sdoc_field] = self._cell_str(ws, row, col_resolver(excel_col))

        return Requirement(
            uid=uid,
            title=title,
            statement=statement,
            custom_fields=custom_fields,
            relations=relations,
        )

    def _write_requirement_to_row(
        self,
        ws: "Worksheet",
        row: int,
        headers: dict[str, int],
        mapping: "SectionMapping",
        req: Requirement,
    ) -> None:
        """Write a single requirement to the given row."""
        if mapping.mode == "table":
            col_resolver = lambda name: headers.get(name)
        else:
            col_resolver = lambda name: column_index_from_string(name) if name else None

        def _set(col_name: str, value: str) -> None:
            cidx = col_resolver(col_name)
            if cidx is not None:
                ws.cell(row, cidx, value)

        _set(mapping.uid_col, req.uid)
        if mapping.title_col:
            _set(mapping.title_col, req.title)
        if mapping.statement_col:
            _set(mapping.statement_col, req.statement)
        if mapping.relations_col:
            _set(mapping.relations_col, "; ".join(req.relations))
        for excel_col, sdoc_field in mapping.extra_cols.items():
            value = req.custom_fields.get(sdoc_field, "")
            if mapping.mode == "table":
                cidx = headers.get(excel_col)
            else:
                cidx = column_index_from_string(excel_col) if excel_col else None
            if cidx is not None:
                ws.cell(row, cidx, value)

    # ------------------------------------------------------------------
    # Lookup helpers
    # ------------------------------------------------------------------

    def _find_table(self, table_name: str) -> tuple["Worksheet", "Table"]:
        """Locate a named table across all sheets."""
        for ws in self._wb.worksheets:  # type: ignore[union-attr]
            for tbl in ws.tables.values():
                if (tbl.displayName or tbl.name) == table_name:
                    return ws, tbl
        raise ExcelEngineError(
            f"Table '{table_name}' not found in workbook '{self.excel_path.name}'."
        )

    def _require_sheet(self, sheet_name: str) -> "Worksheet":
        if sheet_name not in self._wb.sheetnames:  # type: ignore[union-attr]
            available = ", ".join(self._wb.sheetnames)  # type: ignore[union-attr]
            raise ExcelEngineError(
                f"Sheet '{sheet_name}' not found. Available sheets: {available}"
            )
        return self._wb[sheet_name]  # type: ignore[index,union-attr]

    def _require_open(self) -> None:
        if self._wb is None:
            raise ExcelEngineError("Workbook is not open. Call open() first.")

    @staticmethod
    def _col_idx_by_header(
        headers: dict[str, int],
        col_name: str,
        section: str,
        field: str,
    ) -> int:
        if col_name not in headers:
            available = ", ".join(headers.keys())
            raise ExcelEngineError(
                f"[{section}] Column '{col_name}' ({field}) not found. "
                f"Available columns: {available}"
            )
        return headers[col_name]

    @staticmethod
    def _cell_str(ws: "Worksheet", row: int, cidx: int | None) -> str:
        if cidx is None:
            return ""
        val = ws.cell(row, cidx).value
        return str(val) if val is not None else ""
