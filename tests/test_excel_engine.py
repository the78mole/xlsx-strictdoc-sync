"""Tests for the Excel engine (table mode and legacy mode)."""

from __future__ import annotations

import pytest
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

from xlsx_strictdoc_sync.config_manager import SectionMapping
from xlsx_strictdoc_sync.excel_engine import ExcelEngine, ExcelEngineError
from xlsx_strictdoc_sync.models import Requirement


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_table_workbook(tmp_path, rows: list[tuple]) -> str:
    """Create a workbook with a formatted table and return its path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ("UID", "Title", "Statement", "Parent")
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(1, col_idx, h)
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row_idx, col_idx, val)

    last_row = 1 + len(rows)
    tbl = Table(displayName="SYS_Reqs", ref=f"A1:D{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
    ws.add_table(tbl)

    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    return str(path)


def _make_legacy_workbook(tmp_path, rows: list[tuple]) -> str:
    """Create a workbook with a plain sheet and return its path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SYS Requirements"

    headers = ("UID", "Title", "Statement", "Parent")
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(1, col_idx, h)
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row_idx, col_idx, val)

    path = tmp_path / "legacy.xlsx"
    wb.save(str(path))
    return str(path)


def _table_mapping(**kwargs) -> SectionMapping:
    defaults = dict(
        name="SYS_REQS",
        sdoc_file="sys.sdoc",
        mode="table",
        anchor="SYS_Reqs",
        uid_col="UID",
        title_col="Title",
        statement_col="Statement",
        relations_col="Parent",
    )
    defaults.update(kwargs)
    return SectionMapping(**defaults)


def _legacy_mapping(**kwargs) -> SectionMapping:
    defaults = dict(
        name="SYS_REQS",
        sdoc_file="sys.sdoc",
        mode="legacy",
        anchor="SYS Requirements",
        uid_col="A",
        title_col="B",
        statement_col="C",
        relations_col="D",
    )
    defaults.update(kwargs)
    return SectionMapping(**defaults)


# ---------------------------------------------------------------------------
# Table-mode tests
# ---------------------------------------------------------------------------

class TestExcelEngineTableMode:
    def test_read_basic(self, tmp_path):
        path = _make_table_workbook(
            tmp_path,
            [("SYS-001", "Boot", "Shall boot.", ""), ("SYS-002", "Log", "Shall log.", "SYS-001")],
        )
        with ExcelEngine(path) as eng:
            reqs = eng.read_requirements(_table_mapping())

        assert len(reqs) == 2
        assert reqs[0].uid == "SYS-001"
        assert reqs[0].title == "Boot"
        assert reqs[0].statement == "Shall boot."
        assert reqs[0].relations == []
        assert reqs[1].uid == "SYS-002"
        assert reqs[1].relations == ["SYS-001"]

    def test_read_skips_empty_uid(self, tmp_path):
        path = _make_table_workbook(
            tmp_path,
            [("SYS-001", "Boot", "Statement", ""), (None, "Empty", "Row", "")],
        )
        with ExcelEngine(path) as eng:
            reqs = eng.read_requirements(_table_mapping())
        assert len(reqs) == 1

    def test_missing_table_raises(self, tmp_path):
        path = _make_table_workbook(tmp_path, [("SYS-001", "T", "S", "")])
        mapping = _table_mapping(anchor="NonExistentTable")
        with ExcelEngine(path) as eng:
            with pytest.raises(ExcelEngineError, match="not found"):
                eng.read_requirements(mapping)

    def test_write_new_rows(self, tmp_path):
        path = _make_table_workbook(tmp_path, [("SYS-001", "Boot", "Old stmt", "")])
        reqs = [
            Requirement(uid="SYS-001", title="Boot", statement="Updated stmt"),
            Requirement(uid="SYS-002", title="Log", statement="New stmt"),
        ]
        with ExcelEngine(path) as eng:
            eng.write_requirements(reqs, _table_mapping())
            eng.save()

        with ExcelEngine(path) as eng:
            result = eng.read_requirements(_table_mapping())
        uids = [r.uid for r in result]
        assert "SYS-001" in uids
        assert "SYS-002" in uids

    def test_write_updates_existing(self, tmp_path):
        path = _make_table_workbook(tmp_path, [("SYS-001", "Boot", "Old stmt", "")])
        reqs = [Requirement(uid="SYS-001", title="Boot", statement="New stmt")]
        with ExcelEngine(path) as eng:
            eng.write_requirements(reqs, _table_mapping())
            eng.save()

        with ExcelEngine(path) as eng:
            result = eng.read_requirements(_table_mapping())
        assert result[0].statement == "New stmt"

    def test_list_tables(self, tmp_path):
        path = _make_table_workbook(tmp_path, [])
        with ExcelEngine(path) as eng:
            tables = eng.list_tables()
        assert "SYS_Reqs" in tables

    def test_get_table_headers(self, tmp_path):
        path = _make_table_workbook(tmp_path, [])
        with ExcelEngine(path) as eng:
            headers = eng.get_table_headers("SYS_Reqs")
        assert headers == ["UID", "Title", "Statement", "Parent"]


# ---------------------------------------------------------------------------
# Legacy-mode tests
# ---------------------------------------------------------------------------

class TestExcelEngineLegacyMode:
    def test_read_basic(self, tmp_path):
        path = _make_legacy_workbook(
            tmp_path,
            [("SYS-001", "Boot", "Shall boot.", None), ("SYS-002", "Log", "Shall log.", "SYS-001")],
        )
        with ExcelEngine(path) as eng:
            reqs = eng.read_requirements(_legacy_mapping())

        assert len(reqs) == 2
        assert reqs[0].uid == "SYS-001"
        assert reqs[0].title == "Boot"
        assert reqs[1].relations == ["SYS-001"]

    def test_read_skips_empty_uid(self, tmp_path):
        path = _make_legacy_workbook(
            tmp_path,
            [("SYS-001", "T", "S", ""), (None, "Empty", "Row", "")],
        )
        with ExcelEngine(path) as eng:
            reqs = eng.read_requirements(_legacy_mapping())
        assert len(reqs) == 1

    def test_missing_sheet_raises(self, tmp_path):
        path = _make_legacy_workbook(tmp_path, [("SYS-001", "T", "S", "")])
        mapping = _legacy_mapping(anchor="NoSuchSheet")
        with ExcelEngine(path) as eng:
            with pytest.raises(ExcelEngineError, match="Sheet"):
                eng.read_requirements(mapping)

    def test_write_and_read_round_trip(self, tmp_path):
        path = _make_legacy_workbook(tmp_path, [])
        reqs = [
            Requirement(uid="SYS-001", title="T1", statement="S1"),
            Requirement(uid="SYS-002", title="T2", statement="S2", relations=["SYS-001"]),
        ]
        with ExcelEngine(path) as eng:
            eng.write_requirements(reqs, _legacy_mapping())
            eng.save()

        with ExcelEngine(path) as eng:
            result = eng.read_requirements(_legacy_mapping())
        assert len(result) == 2
        assert result[1].relations == ["SYS-001"]

    def test_list_sheets(self, tmp_path):
        path = _make_legacy_workbook(tmp_path, [])
        with ExcelEngine(path) as eng:
            sheets = eng.list_sheets()
        assert "SYS Requirements" in sheets


# ---------------------------------------------------------------------------
# Error handling
# ---------------------------------------------------------------------------

class TestExcelEngineErrors:
    def test_file_not_found(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            with ExcelEngine(tmp_path / "nonexistent.xlsx") as eng:
                eng.read_requirements(_table_mapping())

    def test_not_open_raises(self, tmp_path):
        path = _make_table_workbook(tmp_path, [])
        eng = ExcelEngine(path)
        with pytest.raises(ExcelEngineError, match="not open"):
            eng.read_requirements(_table_mapping())
