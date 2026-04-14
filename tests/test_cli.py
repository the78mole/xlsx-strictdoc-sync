"""Integration tests for the CLI (``reqsync`` command)."""

from __future__ import annotations

import openpyxl
import pytest
from openpyxl.worksheet.table import Table, TableStyleInfo

from xlsx_strictdoc_sync.cli import main


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_toml(tmp_path, excel_file: str, direction: str = "excel_to_sdoc",
               field_directions: dict | None = None) -> str:
    fd_lines = ""
    if field_directions:
        fd_lines = f"\n[SYS_REQS.field_directions]\n"
        for k, v in field_directions.items():
            fd_lines += f'{k} = "{v}"\n'

    content = f"""
[global]
excel_file = "{excel_file}"

[SYS_REQS]
sdoc_file = "{tmp_path}/sys.sdoc"
mode = "table"
anchor = "SYS_Reqs"
uid_col = "UID"
title_col = "Title"
statement_col = "Statement"
relations_col = "Parent"
sync_direction = "{direction}"
{fd_lines}
"""
    path = tmp_path / "reqsync.toml"
    path.write_text(content, encoding="utf-8")
    return str(path)


def _make_excel(tmp_path, rows: list[tuple]) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ("UID", "Title", "Statement", "Parent")
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val)
    last = 1 + len(rows)
    tbl = Table(displayName="SYS_Reqs", ref=f"A1:D{last}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
    ws.add_table(tbl)
    path = tmp_path / "reqs.xlsx"
    wb.save(str(path))
    return str(path)


# ---------------------------------------------------------------------------
# sync – excel_to_sdoc
# ---------------------------------------------------------------------------


class TestCLISyncExcelToSdoc:
    def test_creates_sdoc_from_excel(self, tmp_path):
        excel = _make_excel(tmp_path, [("SYS-001", "Boot", "Shall boot.", "")])
        cfg = _make_toml(tmp_path, excel, direction="excel_to_sdoc")
        rc = main(["sync", cfg])
        assert rc == 0
        sdoc_path = tmp_path / "sys.sdoc"
        assert sdoc_path.exists()
        assert "SYS-001" in sdoc_path.read_text()

    def test_dry_run_no_file_created(self, tmp_path):
        excel = _make_excel(tmp_path, [("SYS-001", "Boot", "S", "")])
        cfg = _make_toml(tmp_path, excel)
        rc = main(["sync", cfg, "--dry-run"])
        assert rc == 0
        assert not (tmp_path / "sys.sdoc").exists()

    def test_direction_flag_overrides_config(self, tmp_path):
        """--direction excel_to_sdoc should work even if config says sdoc_to_excel."""
        excel = _make_excel(tmp_path, [("SYS-001", "Boot", "S", "")])
        cfg = _make_toml(tmp_path, excel, direction="sdoc_to_excel")
        rc = main(["sync", cfg, "--direction", "excel_to_sdoc"])
        assert rc == 0
        assert (tmp_path / "sys.sdoc").exists()


# ---------------------------------------------------------------------------
# sync – sdoc_to_excel
# ---------------------------------------------------------------------------


class TestCLISyncSdocToExcel:
    def test_sdoc_req_written_to_excel(self, tmp_path):
        excel = _make_excel(tmp_path, [])  # empty workbook
        # Pre-populate SDoc file
        sdoc_content = """[DOCUMENT]
TITLE: System Reqs

[GRAMMAR]
ELEMENTS:
- TAG: REQUIREMENT
  FIELDS:
  - TITLE: UID
    TYPE: String
    REQUIRED: False
  - TITLE: TITLE
    TYPE: String
    REQUIRED: False
  - TITLE: STATEMENT
    TYPE: String
    REQUIRED: False

[REQUIREMENT]
UID: SYS-099
TITLE: Boot
STATEMENT: Must boot fast.
"""
        sdoc_path = tmp_path / "sys.sdoc"
        sdoc_path.write_text(sdoc_content)
        cfg = _make_toml(tmp_path, excel, direction="sdoc_to_excel")
        rc = main(["sync", cfg])
        assert rc == 0

        import openpyxl as xl
        wb2 = xl.load_workbook(excel)
        ws2 = wb2.active
        uids = [ws2.cell(r, 1).value for r in range(2, ws2.max_row + 1)]
        assert "SYS-099" in uids


# ---------------------------------------------------------------------------
# sync – both (bidirectional)
# ---------------------------------------------------------------------------


class TestCLISyncBoth:
    def test_new_excel_req_added_to_sdoc(self, tmp_path):
        excel = _make_excel(tmp_path, [("SYS-001", "Boot", "Shall boot.", "")])
        cfg = _make_toml(tmp_path, excel, direction="both")
        rc = main(["sync", cfg])
        assert rc == 0
        assert "SYS-001" in (tmp_path / "sys.sdoc").read_text()

    def test_field_direction_respected(self, tmp_path):
        """TITLE = sdoc_to_excel: after sync the SDoc title should stay as-is."""
        # Pre-populate SDoc with SYS-001 having SDoc-only title
        sdoc_content = """[DOCUMENT]
TITLE: System Reqs

[GRAMMAR]
ELEMENTS:
- TAG: REQUIREMENT
  FIELDS:
  - TITLE: UID
    TYPE: String
    REQUIRED: False
  - TITLE: TITLE
    TYPE: String
    REQUIRED: False
  - TITLE: STATEMENT
    TYPE: String
    REQUIRED: False
  RELATIONS:
  - TYPE: Parent

[REQUIREMENT]
UID: SYS-001
TITLE: SDoc Title
STATEMENT: SDoc Stmt
"""
        sdoc_path = tmp_path / "sys.sdoc"
        sdoc_path.write_text(sdoc_content)

        excel = _make_excel(tmp_path, [("SYS-001", "Excel Title", "Excel Stmt", "")])
        cfg = _make_toml(
            tmp_path, excel,
            direction="both",
            field_directions={"TITLE": "sdoc_to_excel"},
        )
        rc = main(["sync", cfg])
        assert rc == 0
        # SDoc title should remain "SDoc Title" (sdoc_to_excel means SDoc keeps it)
        text = sdoc_path.read_text()
        assert "SDoc Title" in text


# ---------------------------------------------------------------------------
# Section filter
# ---------------------------------------------------------------------------


class TestCLISyncSectionFilter:
    def test_unknown_section_fails(self, tmp_path):
        excel = _make_excel(tmp_path, [("X-001", "T", "S", "")])
        cfg = _make_toml(tmp_path, excel)
        rc = main(["sync", cfg, "--section", "NONEXISTENT"])
        assert rc == 1


# ---------------------------------------------------------------------------
# init-config
# ---------------------------------------------------------------------------


class TestCLIInitConfig:
    def test_generates_toml(self, tmp_path):
        excel = _make_excel(tmp_path, [])
        out = tmp_path / "out.toml"
        rc = main(["init-config", excel, "-o", str(out)])
        assert rc == 0
        assert out.exists()
        assert "sync_direction" in out.read_text()

    def test_missing_excel_fails(self, tmp_path):
        rc = main(["init-config", str(tmp_path / "no.xlsx")])
        assert rc == 1


# ---------------------------------------------------------------------------
# generate-grammar
# ---------------------------------------------------------------------------


class TestCLIGenerateGrammar:
    def test_generates_grammar_file(self, tmp_path):
        excel = _make_excel(tmp_path, [])
        toml = _make_toml(tmp_path, excel)
        rc = main(["generate-grammar", toml, "--output-dir", str(tmp_path)])
        assert rc == 0
        grammar_file = tmp_path / "sys_reqs_grammar.sdoc"
        assert grammar_file.exists()
        assert "REQUIREMENT" in grammar_file.read_text()


# ---------------------------------------------------------------------------
# version / no-args
# ---------------------------------------------------------------------------


class TestCLIMisc:
    def test_no_args_returns_zero(self):
        rc = main([])
        assert rc == 0

    def test_version_raises_system_exit(self):
        with pytest.raises(SystemExit) as exc_info:
            main(["--version"])
        assert exc_info.value.code == 0
